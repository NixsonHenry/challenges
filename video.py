import math
import random
import turtle
import tkinter

"""   
for item in Python:
      print(item)

for item in ["Rundy", "Amione", "Nathou"]:
      print(item)
      
for item in range(10):
      print(item)
for item in range(2, 5):
      print(item)
## --------------------------------------------

prices = [10, 20, 30]

total = 0

for price in prices:
      total += price
print("The total is: ", total)
"""   

""" 
for x in range(4):
      for y in range(3):
        print('(', x, ', ', y, ')')
"""

""" 
numbers = [5, 2, 5, 2, 2]

for x_count in numbers:
    print("X" * x_count)
"""

""" 
numbers = [5, 2, 5, 2, 2]

for x_count in numbers:
    output = ''
    for count in range(x_count):
        output += 'X'
    print(output)
"""
"""
numbers = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, ]

for x_count in numbers:
    output = ''
    for count in range(x_count):
        output += 'X'
    print(output)
"""

## Write a program to find the largest number in a list.
"""
numbers = [3, 6, 2, 8, 4, 10]
max = numbers[0]
for number in numbers:
    if number > max:
        max = number
print(max)
""" 

''' 
## Two dimension lists in Python

matrix = [
    [1,  2,  3],
    [4,  5,  6],
    [7,  8,  9]
]

# print(matrix[0][0])
# print(matrix[0][1])
# print(matrix[0][2]) 

# print(matrix[1][0])
# print(matrix[1][1])
# print(matrix[1][2])

for row in matrix:
    for item in row:
        print(item)
'''

""" 
numbers = [5, 2, 1, 5, 7, 4]

numbers.count(5)
numbers.sort()
numbers.reverse()
numbers.insert(1, 3)
numbers.remove(1)
numbers.clear()
numbers.pop()
numbers.append(13)
numbers1 = numbers.list()
numbers2 = numbers.copy()

print(numbers)
"""

""" 
## Write a program to remove the duplicates in a list.
numbers = [1, 2, 4, 4, 6, 3, 4, 6, 1]
uniques = []
for number in numbers:
    if number not in uniques:
        uniques.append(number)
        uniques.sort()
print(uniques)
"""
## Tuples
"""
    Tuples are similar to list. You can use them to store, but unlike list you cannot modiffy them, you cannot new items, you cannot remove existing items. Tuples are immutable. You cannot mutate or change them.   
"""

"""
## We can use two methods here count and index
numbers = (1, 2, 3)
print(numbers[0])
print(numbers[1])
"""

"""
coordinates = (1, 2, 3)
x =  coordinates[0]
y =  coordinates[1]
z =  coordinates[2]

or, 

x, y, z = coordinates
"""

## Dictionnaries in Python. In a dictionnary, we can store a bunch of value pairs. Each key should be unique
## key-value pairs
## Name, Email, Phone: Keys  & Nixson, xondy77@yahoo.fr: values
""" 
dict = {
    "Name":   "Nixson Henry",
    "Email":  "xondy77@yahoo.fr",
    "Phone":  "(509) 4004-6541"
}
"""

"""
customer = {
    "name":         "John Smith",
    "age":           30,
    "is_verified":   True 
}
customer["name"] 
customer.get("name")
customer["name"] = "Jack Smith"
# Add a new key
x = customer["birthday"] = "Jan 1 1980"
print(customer)
customer.key()
customer.value()
"""

## Convert cardinal phone number taking from user in letters
"""
phone = input("Phone: ")
digits_mapping = {
    "0":    "Zero",
    "1":    "One",
    "2":    "Two",
    "3":    "Three",
    "4":    "Four",
    "5":    "Five",
    "6":    "Six",
    "7":    "Seven",
    "8":    "Eight",
    "9":    "Nine"
}
output = ""
for char in phone:
    output += digits_mapping.get(char, "!") + " "
print(output)
"""

"""
## Smiling face
message = input("> ")
words = message.split(' ')
emojis = {
    ":)":       "😊",
    ":(":       "😒"
}

output = ""
for word in words:
    output += emojis.get(word, word) + ' '
print(output)
"""

## Function is a container for a bunch of codes that perfoem a specific tasks.

"""
def personal_info(firstname, lastname):
    print(f'My fistname is {firstname} and {lastname} is my lastname.')

personal_info("Nixson", "Henry")
"""

"""
def emoji_converter(message):
    words = message.split(' ')
    emojis = {
        ":)":       "😊",
        ":(":       "😒"
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + ' '
    return output


message = input("> ")
print(emoji_converter(message))
"""

## How to handle errors in your python program
## Exit code 1: means the program cratch (Error type: ValueError)
## Exit code 0: means success

## Let's create a new type call point

"""
numbers = [1, 2, 3]
point.get_distance
point.move()
point.draw()
point.move()
"""


#  Pascal naming convention--> Point:, EmailClient: . We don't use underscore to name classes, we capitalize the first letter
# on appelle instance d'une classe, un objet avec un comportement et un état, définis par la classe. Le terme instance s'applique aux variables dont le type est une classe.

# We use classes to define new types. a type is a blue print. With this new type, we can create an object. An object is an instance of a class
# A class define a type or a blue print to create an object. So we can have many objects base on that blue print. 
# We say object or instances. This object can have attributes or properties and that attibute belongs to a particular object

# Creating an object:
# We type the name class and call it as a function
# We can store it in a variable

# When we talk about point, we should know where that point is located. To solve this problem, we use a constructor. A constructor is function that get call at the time we create an object. 

"""
class Point:
    # This constructor method is called every time we create a new object
    def __init__(self, x, y):
        # This __init__ method is used to create, construct or initialize an object where self is a reference to the current object. Self reference that object in memory
        #point2.x = 10
        self.x = x
        self.y = y

    def move(self):
        print("move")
    
    def draw(self):
        print("draw ")

# An object call point1 without the constructor method
# point1 = Point()
# point1.draw()

# An other object call point2 with the constructor method
# Each object is a diffent instance of a print class
point2 = Point(13, 15)
point2.x = 11
print(point2.x)
print(point2.y)
""" 

# Define a new type call: Person
#    - person (its person object has name attribute 
#    - talk() method

""" 
class Person:
    
    def __init__(self, name):
        self.name = name
        
    def talk(self):  # Every method in a class should have the parameter self
        print("Hello, I am", self.name)
        print(self.name, " I'm a computer scientist particularly a web developper.")


name1 = Person("Nixson Henry")
# print(name1.name)
name1.talk()

# Create another person object. So, each object is a different instance of a Person class
name2 = Person("Durdona Abdusamikovna")
name2.talk()
""" 
                                                                        
# Inheritance in Python. Inheritance is a make an example for using code.
# In order to Do not Repeat Yourself (DRY) in Dog and Cat classes, we define a new class, a "Parent class" called Mammal, move the walk() method in Mammal class and pass the name of Parent class (Mammal()) as parameter in the children classes.
# So, Dog and Cat classes inherit from Mammal. 
# Now, we can create Dog and Cat Objects

""" 
class Mammal:
    # def __init__(self, name):
    #     self.name = name

    def walk(self):
        print("walk")



class Dog(Mammal):
    def __init__(self, name):
        self.name = name

    def bark(self):
        print("bark")

dog1 = Dog()
dog1.bark()



class Cat(Mammal):
    def __init__(self, name):
        self.name = name

    def be_annoying(self):
        print("annoying")

cat1 = Cat()
cat1.be_annoying()
"""

"""
num = int(input("Enter a Number: "))
print("Binary value is: {}".format(bin(num)))
print("Octal value is: {}".format(oct(num)))
print("Hexadecimmal value is: {}".format(hex(num)))
"""

"""
vowels = ('a', 'e', 'i', 'o', 'u')
message = input("Enter Message: ")
new_message = []

for letters in message:
    if letters not in vowels:
        new_message += letters
print("Message without vowels: {}".format(new_message))
"""

"""
my_dict = {
    "A":"E", "B":"F", "C":"G", "D":"H", "E":"I",
    "F":"J", "G":"K", "H":"L", "I":"M", "J":"N",
    "K":"O", "L":"P", "M":"Q", "N":"R", "O":"S",
    "P":"T", "Q":"U", "R":"V", "S":"W", "T":"X",
    "U":"Y", "V":"Z", "W":"A", "X":"B", "Y":"C",
    "Z":"D", "1":"4", "2":"5", "3":"6", "4":"7",
    "5":"8", "6":"9", "7":"0", "8":"1", "9":"2",
    "0":"3", " ":" "}

message = input("Enter your message: ").upper()
encrypted = ""

for letters in message:
    if letters in my_dict:
        encrypted += my_dict[letters]
    else:
        encrypted += letters
print(encrypted.lower())
"""

"""
guess = []
rand_number = random.randint(1, 100)
player_num = int(input("Enter a number from 1 to 100: "))
guess.append(player_num)
print(rand_number)


while player_num != rand_number:
    if player_num > rand_number:
        print("Your number is too high !!!")
    else:
        print("Your number is too low !!!")
    player_num = int(input("Enter a number from 1 to 100: "))
    guess.append(player_num)
else:
    print("Well done...")
    print("You have taken {} guesses".format(len(guess)))
    print("Here are your guesses: ")
    print(guess)
"""


"""
import sys

def modulo(x, y):
    return(x % y)

def add(x, y):
    return(x + y)

def substract(x, y):
    return(x - y)
    
def multiply(x, y):
    return(x * y)


def divide(x, y):
    try:
        result = (x / y)
        return result
    except ZeroDivisionError:
        # Exit the program
        print("You cannot divide by zero. ")
        # sys.exit()

def power(x, y):
    return(x**y)


def square_root(x):
    try:
        print(f'Square Root of {x} is {math.sqrt(x)} ')
    except ValueError:
        print(f'You entered {x}, which is not a positive number.')
  

print("Select your operation: ")
print("0: Modulo")
print("1: Add")
print("2: Substract")
print("3: Multiply")
print("4: Divide")
print("5: Power")
print("6: Square root")


try:
    choice = input("Enter your choice: 0/1/2/3/4/5/6: ")
    num1 = float(input("Enter the first number: "))
    num2 = float(input("Enter the second number: "))
except TypeError:
    print('Invalid value')
    sys.exit()

if choice == '0':
    print(num1, " % ", num2, " = ", modulo(num1, num2))
elif choice == '1':
    print(num1, " + ", num2, " = ", add(num1, num2))
elif choice == '2':
    print(num1, " - ", num2, " = ", substract(num1, num2))
elif choice == '3':
    print(num1, " * ", num2, " = ", multiply(num1, num2))
elif choice == '4':
    print(num1, " / ", num2, " = ", divide(num1, num2))
elif choice == '5':
    print(num1, " ** ", num2, " = ", power(num1, num2))
elif choice == '6':
    square_root(num1)
else:
    print("Invalid input")

"""



""" 
jours = ['Dimache', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']

a = int(input('Enter the minimum value for your range: '))
b = int(input('Enter the maximun value for your range: '))

for i in range(a, b):
    print(jours[i])
""" 
    
""" 
questions = {
    'question1':    "What is your name? ",
    'question2':    "What old are you? ",
    'question3':    "What is your profession? ",
    'question4':    "What do you do on your favorite time? ",
    'question5':    "Where do you live? ",
    'question6':    "Where are you from? ",
    'question7':    "What is your mother's name? ",
    'question8':    "How many brothers and sisters are you? ",
    'question9':    "What is your favorite movie? ",
    'question10':   "Are you christian? "
}

# new_list = []
# n = random.randint(0, 9) 
# print(n)
# questions_list = questions.values()

# for question in questions_list:
#     new_list.append(question)
# print(new_list[n])

# print(random.choice([])) 

# 2eme facon
# ke = list(questions.keys())
# ke2 = (random.choice(ke))


# print(questions[ke2]) 

a = tkinter._test()
b = tkinter.Tcl().eval('info patchlevel')
""" 

""" 
turtle.shape("turtle")

for i in range(0, 5):
    turtle.forward(100)
    turtle.right(72)

turtle.exitonclick()
""" 

# Challenge 80 a 117

# Let's take these functions and put them in a seperate module called "Converters" and import that module in any program that needs these converter function

# Convert kg to lbs and lbs into kg using the converters module 
""" 
import converters

weights = converters.kg_to_lbs(63)

print("You weigh: ", weights, 'lbs')
""" 
# Find the largest number in a list using the utils module 

"""
def find_max(list):
    max = list[0]
    for number in list_2:
        if number > max:
            max = number
    return max


def numbers():
    n = int(input("length of list: "))
    numbers = []
    for i in range(n):
        element = input("Inter the element: ")
        numbers.append(element) 
    return numbers 

list_2 = numbers()
maximum = find_max(list_2)
print(maximum)


# print(f"The maximum number of the following list {numbers} is: {maximum} ")
"""


# Package is a container for multiple modules. In file system terms, package is a directory or folder
# Let's create a package called ecommerce with a python file that identifies the package, and module called shipping.py 
# Package for: Men's -------- Women's --------- Kids'
         # a)  Shoes
         # b)  T-Shirt,  we call a), b) and c): A MODULE
         # c>  Jackets
# We have to import this shipping module into our app.py module: 


"""
    
import random

# members = ['john', 'Mary', 'Bob', 'Mosh']
# leader = random.choice(members)
# print(leader)

# A rolling dice

# for i in range(2):
#     face = [1, 2, 3, 4, 5, 6]
#     rand = random.choice(face)
# print(rand)

# for i in range(3):
    # print(random.random())
    # print(random.randint(10, 20))


class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second


dice = Dice()
print(dice.roll())

# How to work with directories in Python

"""

# from pathlib import Path

# Absolute path: To the route
# C:\Program File\Microsoft   in windows
# /user/local/bin in Mac or Linux
# Relative path: when we start with the current directory, in the package (ecommerce) to somewhere else:
# path = Path("eccommerce")

# To work with excel files


"""
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    # wb = xl.load_workbook("transactions.xlsx")
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']  # Gi me the coordinates of that sheet or
    cell = sheet.cell(1, 1) # Gi me the coordinates of that sheet
    # print(cell.value)

    # How many rows we have in this spreadsheet?
    # print(sheet.max_row)
    # for row in range(1, sheet.max_row + 1):
        # print(row)
    for row in range(2, sheet.max_row + 1): # Since we don't need the heading, the first row
        cell = sheet.cell(row, 3)
        # print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell = corrected_price 
    values = Reference(sheet,
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'g2')
    wb.save('transactions2.xlsx')
    wb.save(filename)
""" 

""" 
class Flight():

    def __init__(self, capacity):
        self.capacity = capacity
        self.passengers = []

    
    # Add passenger to the flight
    def add_passenger(self, name):
        if not self.open_seats():
            return False
        else:
            self.passengers.append(name)
            return True


    # To return the number of seats
    def open_seats(self):
        return self.capacity - len(self.passengers)



# Create a new flight with o=up to 7 passengers
flight = Flight(7)

# Create a list of people
people = ["Harry", "Rundy", "Hermione", "Ginny", "Syndie", "Dona", "Nathanel", "Susan", "Dave"]

# Attempt to add each person in the list to a flight
for person in people:
    if flight.add_passenger(person):
         print(f"Added {person} to flight successfully")
    else:
        print(f"No available seats for {person}")
""" 


"""
# https://www.journaldev.com/14633/python-inheritance-example
#Line:1, definition of the superclass starts here  
class Person:  
    #initializing the variables  
    name = ""  
    age = 0  
  
    #defining constructor  
    def __init__(self, personName, personAge):  
        self.name = personName  
        self.age = personAge  
  
    #defining class methods  
    def showName(self):  
        print(self.name)  
  
    def showAge(self):  
        print(self.age)  
  
    #Line: 19, end of superclass definition  
  
#definition of subclass starts here
class Student(Person): #Line: 22, Person is the  superclass and Student is the subclass
    studentId = ""  
  
    def __init__(self, studentName, studentAge, studentId):  
        Person.__init__(self, studentName, studentAge)  #Line: 26, Calling the superclass constructor and sending values of attributes.
        self.studentId = studentId  
  
    def getId(self):  
        return self.studentId  #returns the value of student id
#end of subclass definition
  
  
# Create an object of the superclass  
person1 = Person("Richard", 23)  #Line: 35
#call member methods of the objects  
person1.showAge()  
# Create an object of the subclass  
student1 = Student("Max", 22, "102")  #Line: 39
print(student1.getId())  
student1.showName()  #Line: 41
""" 

""" 
class A:
    def __init__(self):
        print('Initializing: class A')

    def sub_method(self, b):
        print('Printing from class A:', b)


class B(A):
    def __init__(self):
        print('Initializing: class B')
        super().__init__()

    def sub_method(self, b):
        print('Printing from class B:', b)
        super().sub_method(b + 1)


class C(B):
    def __init__(self):
        print('Initializing: class C')
        super().__init__()

    def sub_method(self, b):
        print('Printing from class C:', b)
        super().sub_method(b + 1)


if __name__ == '__main__':
    c = C()
    c.sub_method(1)
""" 


"""
#definition of the class starts here  
class Person:  
    #defining constructor  
    def __init__(self, personName, personAge):  
        self.name = personName  
        self.age = personAge  
  
    #defining class methods  
    def showName(self):  
        print(self.name)  
  
    def showAge(self):  
        print(self.age)  
  
    #end of class definition  
  
# defining another class  
class Student: # Person is the  
    def __init__(self, studentId):  
        self.studentId = studentId  
  
    def getId(self):  
        return self.studentId  
  
  
class Resident(Person, Student): # extends both Person and Student class  
    def __init__(self, name, age, id):  
        Person.__init__(self, name, age)  
        Student.__init__(self, id)  
  
  
# Create an object of the subclass  
resident1 = Resident('John', 30, '102')  
resident1.showName()  
print(resident1.getId())  
"""

"""
class A:  
    def __init__(self):  
        self.name = 'John'  
        self.age = 23  
  
    def getName(self):  
        return self.name  
  
  
class B:  
    def __init__(self):  
        self.name = 'Richard'  
        self.id = '32'  
  
    def getName(self):  
        return self.name  
  
  
class C(A, B):  
    def __init__(self):  
        A.__init__(self)  
        B.__init__(self)  
  
    def getName(self):  
        return self.name  
  
C1 = C()  
print(C1.getName())  
"""


"""
class A:  
    def __init__(self):  
        super().__init__()  
        self.name = 'John'  
        self.age = 23  
  
    def getName(self):  
        return self.name  
  
  
class B:  
    def __init__(self):  
        super().__init__()  
        self.name = 'Richard'  
        self.id = '32'  
  
    def getName(self):  
        return self.name  
  
  
class C(A, B):  
    def __init__(self):  
        super().__init__()  
  
    def getName(self):  
        return self.name  
  
C1 = C()  
print(C1.getName())  
"""


"""
# Python program showing a use
# of get() and set() method in
# normal function
  
class Geek:
    def __init__(self, age = 0):
         self._age = age
      
    # getter method
    def get_age(self):
        return self._age
      
    # setter method
    def set_age(self, x):
        self._age = x
  
raj = Geek()
  
# setting the age using setter
raj.set_age(21)
  
# retrieving age using getter
print(raj.get_age())
  
print(raj._age)
"""



"""
# Python program showing a
# use of property() function
  
class Geeks:
     def __init__(self):
          self._age = 0
       
     # function to get value of _age
     def get_age(self):
         print("getter method called")
         return self._age
       
     # function to set value of _age
     def set_age(self, a):
         print("setter method called")
         self._age = a
  
     # function to delete _age attribute
     def del_age(self):
         del self._age
     
     age = property(get_age, set_age, del_age) 
  
mark = Geeks()
  
mark.age = 10
  
print(mark.age)
"""



"""
# https://www.datacamp.com/community/tutorials/property-getters-setters?utm_source=adwords_ppc&utm_campaignid=1455363063&utm_adgroupid=65083631748&utm_device=c&utm_keyword=&utm_matchtype=b&utm_network=g&utm_adpostion=&utm_creative=278443377095&utm_targetid=dsa-429603003980&utm_loc_interest_ms=&utm_loc_physical_ms=1007622&gclid=Cj0KCQjwg7KJBhDyARIsAHrAXaFCXYPJEGgH95jKjbpitKvsys4OBp12oZOu9cQM-9NeosCXQtbXQ8UaApwMEALw_wcB
class AnotherWay:
    
    def __init__(self, var):
        ## calling the set_a() method to set the value 'a' by checking certain conditions
        self.set_a(var)

    ## getter method to get the properties using an object
    def get_a(self):
        return self.__a

    ## setter method to change the value 'a' using an object
    def set_a
    (self, var):

        ## condition to check whether var is suitable or not
        if var > 0 and var % 2 == 0:
            self.__a = var
        else:
            self.__a = 2

    a = property(get_a, set_a)
"""



"""
class FinalClass:
    
    def __init__(self, var):
        ## calling the set_a() method to set the value 'a' by checking certain conditions
        self.__set_a(var)

    ## getter method to get the properties using an object
    def __get_a(self):
        return self.__a

    ## setter method to change the value 'a' using an object
    def __set_a(self, var):

        ## condition to check whether var is suitable or not
        if var > 0 and var % 2 == 0:
            self.__a = var
        else:
            self.__a = 2

    a = property(__get_a, __set_a)
"""

